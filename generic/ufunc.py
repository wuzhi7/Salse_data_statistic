import pandas as pd
import os


# 定义一个函数，上下合并表格
def v_concat_tables(folder) -> pd.DataFrame:
    df_list = []  # 一个空的数据列表
    for fn in os.listdir(folder):  # 遍历一个文件夹中的所有文件

        afn = os.path.join(folder, fn)  # 定义一个绝对路径
        #         print(afn)
        df_temp = pd.read_excel(afn)  # 读取每一个Excel文件
        df_list.append(df_temp)  # 把每一个读取的文件添加到空的列表中
    df = pd.concat(df_list, ignore_index=True)  # 上下拼接所有列表，形成一个dataframe
    # print(df)
    return df

# 插入图片函数
from openpyxl import load_workbook
from openpyxl.drawing.image import Image



def insert_img_to_excel(filename, by_col, to_col, img_folder):
    """
    插入图片到excel
    :param filename:excel文件路径
    :param by_col:选择哪一列
    :param to_col:插入到哪一列
    :param img_folder:图片源文件夹
    :return: None
    """
    wb = load_workbook(filename)
    ws = wb.active
    for i, c in enumerate(ws[by_col], start=1):
        #         print(i, c.value)
        img_ffn = os.path.join(img_folder, c.value + '.jpg')  # 根据列名生成一个相应的图片路径
        print(i, img_ffn)
        try:
            img = Image(img_ffn)
            img.height = 80
            img.width = 80
            ws.add_image(img=img, anchor=to_col + str(i))  # str(i) 把int类型转化为str anchor根据前一列进行匹配
        except:
            print(c.value, '匹配不到图片')
    wb.save(filename)








if __name__ == '__main__':
    df_jd = v_concat_tables('/Users/liyangbin/PycharmProjects/Sales/data_source/jd_sales')
    df_jd.to_excel('jd_sales.xlsx')
    # insert_img_to_excel('1.xlsx', 'A', 'B', img_folder='img')
